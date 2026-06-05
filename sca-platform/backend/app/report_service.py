from __future__ import annotations

import html
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .license_policy import license_policy, normalize_license_name
from .models import Component, Project, RiskMonitorSnapshot, ScanTask, UploadFileRecord, VulnerabilityRecord


CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
</Types>"""

WORD_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""

XLSX_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>"""

XLSX_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""

WORKBOOK = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="安全分析报告" sheetId="1" r:id="rId1"/></sheets></workbook>"""

WORKBOOK_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>"""

WORD_STYLES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:docDefaults><w:rPrDefault><w:rPr><w:rFonts w:ascii="Calibri" w:eastAsia="宋体" w:hAnsi="Calibri"/><w:sz w:val="21"/><w:color w:val="1F2937"/></w:rPr></w:rPrDefault></w:docDefaults>
<w:style w:type="paragraph" w:default="1" w:styleId="Normal"><w:name w:val="Normal"/><w:pPr><w:spacing w:after="120" w:line="360" w:lineRule="auto"/></w:pPr></w:style>
<w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/><w:pPr><w:jc w:val="center"/></w:pPr><w:rPr><w:rFonts w:eastAsia="黑体"/><w:b/><w:sz w:val="42"/><w:color w:val="111827"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Subtitle"><w:name w:val="Subtitle"/><w:pPr><w:jc w:val="center"/></w:pPr><w:rPr><w:rFonts w:eastAsia="宋体"/><w:sz w:val="24"/><w:color w:val="374151"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/><w:basedOn w:val="Normal"/><w:pPr><w:keepNext/><w:spacing w:before="360" w:after="180"/><w:outlineLvl w:val="0"/></w:pPr><w:rPr><w:rFonts w:eastAsia="黑体"/><w:b/><w:sz w:val="30"/><w:color w:val="0F172A"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Heading2"><w:name w:val="heading 2"/><w:basedOn w:val="Normal"/><w:pPr><w:keepNext/><w:spacing w:before="240" w:after="120"/><w:outlineLvl w:val="1"/></w:pPr><w:rPr><w:rFonts w:eastAsia="黑体"/><w:b/><w:sz w:val="24"/><w:color w:val="1F2937"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="TOC"><w:name w:val="toc"/><w:pPr><w:spacing w:after="60"/></w:pPr><w:rPr><w:sz w:val="21"/></w:rPr></w:style>
</w:styles>"""

SEVERITY_LABELS = {
    "critical": "严重风险",
    "high": "高危风险",
    "medium": "中危风险",
    "low": "低危风险",
    "unknown": "未知风险",
}

COMPONENT_RISK_RANK = {"严重风险": 4, "高危风险": 3, "中危风险": 2, "低危风险": 1, "未知风险": 0, "无漏洞": -1}
UNKNOWN_TEXT_VALUES = {"", "unknown", "none", "null", "n/a", "na", "待确认", "未知"}


def _is_unknown_text(value: object) -> bool:
    return str(value or "").strip().lower() in UNKNOWN_TEXT_VALUES


def _display_license(value: object) -> str:
    text = normalize_license_name(str(value or "").strip())
    return "未声明" if _is_unknown_text(text) else text


def _license_source_label(value: object) -> str:
    source = str(value or "").strip()
    mapping = {
        "npm_registry": "npm 注册表",
        "pypi_registry": "PyPI 注册表",
        "maven_pom": "Maven POM",
        "package-lock.json": "package-lock.json",
        "composer.lock": "composer.lock",
        "METADATA": "Python METADATA",
    }
    return mapping.get(source, source or "未识别")


def _component_examples(names: list[str], limit: int = 12) -> str:
    unique = sorted(dict.fromkeys(name for name in names if name))
    if len(unique) <= limit:
        return "、".join(unique)
    return "、".join(unique[:limit]) + f" 等 {len(unique)} 个"


def _license_confidence(values: list[float]) -> str:
    usable = [float(value) for value in values if value]
    if not usable:
        return "需人工确认"
    return f"{round(sum(usable) / len(usable) * 100)}%"


def _license_rows(components: list[Component]) -> list[list[object]]:
    grouped: dict[str, dict[str, object]] = {}
    for component in components:
        name = _display_license(component.license_name)
        item = grouped.setdefault(name, {"components": [], "sources": [], "confidence": []})
        item["components"].append(component.package_name)
        item["sources"].append(component.license_source)
        item["confidence"].append(component.license_confidence)
    rows: list[list[object]] = []
    for name in sorted(grouped, key=lambda value: (value == "未声明", value)):
        policy = license_policy(name)
        item = grouped[name]
        sources = sorted(dict.fromkeys(_license_source_label(source) for source in item["sources"]))
        rows.append(
            [
                policy.short_name,
                policy.full_name,
                policy.risk_note,
                policy.scope,
                policy.conditions,
                policy.limitations,
                policy.gpl_compatible,
                policy.osi_approved,
                policy.fsf_approved,
                policy.risk_level,
                policy.description,
                len(item["components"]),
                _component_examples(item["components"]),
                "、".join(sources) if sources else "未识别",
                _license_confidence(item["confidence"]),
            ]
        )
    return rows


def _looks_like_version_range(value: object) -> bool:
    text = str(value or "").strip().lower()
    return bool(
        text.startswith(("^", "~", ">", "<", ">=", "<=", "~=", "[", "("))
        or "," in text
        or "*" in text
        or text.endswith(".+")
        or " - " in text
        or text in {"latest", "latest.release", "latest.integration", "snapshot"}
        or text.endswith("-snapshot")
    )


def _snapshot_inferred_version(component: Component, snapshot: RiskMonitorSnapshot | None) -> str:
    if not snapshot:
        return ""
    current = str(component.package_version or "").strip()
    snapshot_version = str(snapshot.current_version or "").strip()
    if (_is_unknown_text(current) or not component.version_detected) and not _is_unknown_text(snapshot_version):
        return snapshot_version
    return ""


def _component_unknown_version_count(components: list[Component], version_cache: dict[int, RiskMonitorSnapshot] | None = None) -> int:
    version_cache = version_cache or {}
    return sum(
        1
        for component in components
        if (_is_unknown_text(component.package_version) or not component.version_detected)
        and not _snapshot_inferred_version(component, version_cache.get(component.id))
    )


def _display_component_version(component: Component, snapshot: RiskMonitorSnapshot | None = None) -> str:
    inferred_version = _snapshot_inferred_version(component, snapshot)
    if inferred_version:
        return f"{inferred_version}（未声明，按默认安装推断）"
    current = str(component.package_version or "").strip()
    if _is_unknown_text(current) or not component.version_detected:
        return "未声明版本"
    declared = str(component.declared_version or "").strip()
    normalized = str(component.version_normalized or "").strip()
    if _looks_like_version_range(current) and not _is_unknown_text(normalized) and normalized != current:
        return f"{normalized}（声明：{current}）"
    if declared and declared != current and _looks_like_version_range(declared):
        return f"{current}（声明：{declared}）"
    return current


def _vulnerability_component_version(component: Component | None, vulnerability: VulnerabilityRecord) -> str:
    if component:
        display_version = _display_component_version(component)
        if display_version != "未声明版本":
            return display_version
    vulnerability_version = str(vulnerability.package_version or "").strip()
    if not _is_unknown_text(vulnerability_version):
        return vulnerability_version
    if component:
        for candidate in (component.resolved_version, component.version_normalized, component.package_version):
            text = str(candidate or "").strip()
            if not _is_unknown_text(text):
                return text
    return "未声明版本"


def _project_data(db: Session, project_id: int) -> tuple[Project, list[Component], list[VulnerabilityRecord]]:
    project = db.get(Project, project_id)
    if not project:
        raise ValueError("项目不存在")
    components = list(db.scalars(select(Component).where(Component.project_id == project_id).order_by(Component.ecosystem, Component.package_name)))
    vulnerabilities = list(db.scalars(select(VulnerabilityRecord).where(VulnerabilityRecord.project_id == project_id)))
    return project, components, vulnerabilities


def _report_metadata(metadata: dict[str, object] | None = None) -> dict[str, str]:
    today = datetime.now(timezone.utc).strftime("%Y.%m.%d")
    defaults = {
        "client_name": "XXXX公司",
        "client_address": "",
        "contact_name": "",
        "contact_phone": "",
        "contact_email": "",
        "organization_name": "XXXXXX有限公司",
        "audit_address": "",
        "auditor_name": "平台自动扫描",
        "reviewer_name": "安全分析人员",
        "quality_reviewer_name": "质量审核人员",
        "accepted_date": today,
        "audit_start_date": today,
        "audit_end_date": today,
        "version_number": "V1.0",
    }
    for key, value in (metadata or {}).items():
        if key in defaults and value not in {None, ""}:
            defaults[key] = str(value)
    return defaults


def _latest_version_cache(db: Session, project_id: int) -> dict[int, RiskMonitorSnapshot]:
    snapshots = list(db.scalars(select(RiskMonitorSnapshot).where(RiskMonitorSnapshot.project_id == project_id)))
    snapshots.sort(key=lambda item: ((item.checked_at.isoformat() if item.checked_at else ""), item.id or 0), reverse=True)
    latest: dict[int, RiskMonitorSnapshot] = {}
    for snapshot in snapshots:
        if snapshot.component_id and snapshot.component_id not in latest:
            latest[snapshot.component_id] = snapshot
    return latest


def _upload_records(db: Session, project_id: int) -> list[UploadFileRecord]:
    return list(db.scalars(select(UploadFileRecord).where(UploadFileRecord.project_id == project_id).order_by(UploadFileRecord.created_at.desc())))


def _risk_counts(vulnerabilities: list[VulnerabilityRecord]) -> dict[str, int]:
    counts = {"critical": 0, "high": 0, "medium": 0, "low": 0, "unknown": 0}
    for vulnerability in _confirmed_vulnerabilities(vulnerabilities):
        counts[vulnerability.severity if vulnerability.severity in counts else "unknown"] += 1
    return counts


def _confirmed_vulnerabilities(vulnerabilities: list[VulnerabilityRecord]) -> list[VulnerabilityRecord]:
    return [item for item in vulnerabilities if item.match_status == "affected" and not item.needs_human_review]


def _priority_rank(value: str) -> int:
    return {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "Review": 4, "Ignore": 5}.get(value or "Review", 4)


def _priority_sorted(vulnerabilities: list[VulnerabilityRecord]) -> list[VulnerabilityRecord]:
    return sorted(vulnerabilities, key=lambda item: (_priority_rank(item.risk_priority), -float(item.risk_score or 0), -float(item.cvss_score or 0)))


def _vulnerability_review_status(item: VulnerabilityRecord) -> str:
    if item.risk_priority == "Ignore" or item.false_positive_possibility == "high":
        return "疑似误报"
    if item.needs_human_review or item.match_status == "unknown" or item.risk_priority == "Review":
        return "待确认"
    return "已确认"


def _confidence_groups(vulnerabilities: list[VulnerabilityRecord]) -> dict[str, int]:
    groups = {"confirmed": 0, "review": 0, "false_positive": 0}
    for item in vulnerabilities:
        status = _vulnerability_review_status(item)
        if status == "疑似误报":
            groups["false_positive"] += 1
        elif status == "待确认":
            groups["review"] += 1
        else:
            groups["confirmed"] += 1
    return groups


def _component_confidence_groups(components: list[Component], version_cache: dict[int, RiskMonitorSnapshot] | None = None) -> dict[str, int]:
    return {
        "high": sum(1 for item in components if item.confidence_level == "High"),
        "medium": sum(1 for item in components if item.confidence_level in {"Medium", "Medium-High"}),
        "low": sum(1 for item in components if item.confidence_level == "Low"),
        "review": sum(1 for item in components if item.confidence_level == "Review"),
        "unknown_version": _component_unknown_version_count(components, version_cache),
        "manual_confirm": sum(1 for item in components if item.need_manual_confirm or item.need_manual_version_confirm),
    }


def _status_label(status: str) -> str:
    return {
        "completed": "成功",
        "success": "成功",
        "failed": "失败",
        "timeout": "超时",
        "skipped": "跳过",
        "partial_completed": "部分完成",
        "running": "执行中",
        "pending": "等待中",
    }.get(status, status or "未知")


def _scan_tool_details(tasks: list[ScanTask]) -> list[dict[str, str]]:
    parent_ids = [item.id for item in tasks if item.parent_task_id is None]
    latest_parent_id = max(parent_ids) if parent_ids else None
    detail_tasks = [
        item
        for item in tasks
        if item.engine_name in {"opensca", "syft", "trivy", "dependency-track"}
        and (latest_parent_id is None or item.parent_task_id == latest_parent_id)
    ]
    return [
        {
            "engine": item.engine_name,
            "status": item.status,
            "status_label": _status_label(item.status),
            "reason": item.error_message or item.summary or "",
            "report_path": item.raw_result_path or item.normalized_result_path or "",
        }
        for item in sorted(detail_tasks, key=lambda row: row.id)
    ]


def _tool_status_text(details: object) -> str:
    if not isinstance(details, list) or not details:
        return "暂无工具状态明细。"
    lines = []
    for item in details:
        if not isinstance(item, dict):
            continue
        reason = str(item.get("reason") or "无")
        report_path = str(item.get("report_path") or "无")
        lines.append(f"{item.get('engine')}：{item.get('status_label')}；原因/提示：{reason}；报告文件：{report_path}")
    return "；".join(lines) if lines else "暂无工具状态明细。"


def _tool_status_rows(scan_tasks: dict[str, object]) -> list[list[object]]:
    rows: list[list[object]] = [["扫描引擎", "状态", "失败原因/提示", "报告文件路径"]]
    details = scan_tasks.get("details")
    if isinstance(details, list):
        for item in details:
            if not isinstance(item, dict):
                continue
            rows.append(
                [
                    item.get("engine") or "-",
                    item.get("status_label") or item.get("status") or "-",
                    item.get("reason") or "无",
                    item.get("report_path") or "无",
                ]
            )
    if len(rows) == 1:
        rows.append(["-", "暂无", "暂无工具状态明细", "无"])
    return rows


def _scan_task_summary(db: Session, project_id: int) -> dict[str, object]:
    tasks = list(db.scalars(select(ScanTask).where(ScanTask.project_id == project_id)))
    return {
        "opensca": sum(1 for item in tasks if item.engine_name == "opensca" and item.status in {"completed", "success"}),
        "syft": sum(1 for item in tasks if item.engine_name == "syft" and item.status in {"completed", "success"}),
        "trivy": sum(1 for item in tasks if item.engine_name == "trivy" and item.status in {"completed", "success"}),
        "dependency_track": sum(1 for item in tasks if item.engine_name == "dependency-track" and item.status in {"completed", "success"}),
        "failed": sum(1 for item in tasks if item.status in {"failed", "timeout"}),
        "skipped": sum(1 for item in tasks if item.status == "skipped"),
        "partial": sum(1 for item in tasks if item.status == "partial_completed"),
        "details": _scan_tool_details(tasks),
    }


def _release_advice(vulnerabilities: list[VulnerabilityRecord]) -> str:
    priorities = {item.risk_priority for item in vulnerabilities}
    if "P0" in priorities:
        return "不建议上线：存在 P0 级风险，需立即整改并复测。"
    if "P1" in priorities:
        return "建议整改后上线：存在 P1 级风险，应完成升级或缓解措施后再发布。"
    if priorities & {"Review"}:
        return "谨慎上线：存在待确认漏洞，建议完成证据复核后再进入生产发布。"
    return "可按流程上线：未发现需要阻断发布的已确认高优先级漏洞。"


def _natural_summary(project: Project, components: list[Component], vulnerabilities: list[VulnerabilityRecord]) -> str:
    confirmed = _confirmed_vulnerabilities(vulnerabilities)
    priority_counts = {level: sum(1 for item in confirmed if item.risk_priority == level) for level in ("P0", "P1", "P2", "P3")}
    risky_components = ", ".join(dict.fromkeys(item.package_name for item in _priority_sorted(confirmed)[:5])) or "暂无集中风险组件"
    urgent = priority_counts["P0"] + priority_counts["P1"]
    if urgent:
        risk_text = f"本次扫描发现 {urgent} 个需要优先整改的 P0/P1 风险"
    else:
        risk_text = "本次扫描未发现需要立即阻断的 P0/P1 已确认风险"
    return f"{risk_text}，风险主要集中在 {risky_components}。{_release_advice(confirmed)} 项目备注：{project.scan_note or '无'}。组件总数 {len(components)}，漏洞记录 {len(vulnerabilities)}。"


def _fix_command(component: Component | None, vulnerability: VulnerabilityRecord) -> str:
    name = vulnerability.package_name
    fixed = vulnerability.fixed_version or "安全版本"
    ecosystem = (vulnerability.ecosystem or getattr(component, "ecosystem", "") or "").lower()
    if ecosystem in {"maven", "java"}:
        return f"Maven：将 {name} 的版本升级到 {fixed}，并执行 mvn dependency:tree 验证传递依赖。"
    if ecosystem in {"npm", "node", "javascript"}:
        return f"npm：优先修改 package.json/lockfile 到 {name}@{fixed}，再执行 npm install 与 npm audit 验证。"
    if ecosystem in {"pypi", "python"}:
        return f"pip：执行 pip install {name}=={fixed}，并同步 requirements.txt/poetry.lock。"
    if ecosystem in {"go", "golang"}:
        return f"Go：执行 go get {name}@{fixed}，再运行 go mod tidy。"
    if ecosystem in {"docker", "container"}:
        return f"Docker：将基础镜像或镜像组件 {name} 升级到 {fixed}，重新构建并扫描镜像。"
    return f"通用：将 {name} 升级到 {fixed}，升级后重新扫描确认漏洞消除。"


def _report_lines(
    project: Project,
    components: list[Component],
    vulnerabilities: list[VulnerabilityRecord],
    scan_tasks: dict[str, object] | None = None,
    version_cache: dict[int, RiskMonitorSnapshot] | None = None,
) -> list[str]:
    counts = _risk_counts(vulnerabilities)
    confirmed = _confirmed_vulnerabilities(vulnerabilities)
    high_risk = [item for item in confirmed if item.severity in {"critical", "high"}]
    confidence = _confidence_groups(vulnerabilities)
    component_confidence = _component_confidence_groups(components, version_cache)
    scan_tasks = scan_tasks or {}
    component_by_id = {item.id: item for item in components}
    priority_items = _priority_sorted(confirmed)
    fallback_enabled = any(item.detected_by == "fallback" for item in components)
    modes = sorted({item.scan_mode for item in components if item.scan_mode})
    unlocked_components = [item for item in components if item.version_risk_type]
    priority_lines = [
        f"{item.risk_priority or 'Review'}：{item.cve_id or item.advisory_id} / {item.package_name} {item.package_version}，建议期限 {item.suggested_deadline}"
        for item in priority_items[:12]
    ]
    fix_lines = [_fix_command(component_by_id.get(item.component_id or 0), item) for item in priority_items[:12]]
    unlocked_lines = [
        f"{item.ecosystem} / {item.source_file or item.source_path} / {item.package_name} / 声明版本 {item.declared_version or '未声明'} / 实际版本 {item.resolved_version or item.package_version} / {item.version_risk_type or item.version_lock_status} / {item.risk_explanation}"
        for item in unlocked_components[:20]
    ]
    lines = [
        "聚信软件成分安全分析报告",
        "企业 Logo：JUXIN",
        "一、本次扫描结论摘要",
        _natural_summary(project, components, vulnerabilities),
        f"上线建议：{_release_advice(confirmed)}",
        "是否需要立即整改：" + ("是，存在 P0/P1 风险。" if any(item.risk_priority in {"P0", "P1"} for item in confirmed) else "否，建议按常规整改计划推进。"),
        f"项目概况：{project.name}",
        f"扫描时间：{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
        f"扫描备注：{project.scan_note}",
        f"组件统计：共 {len(components)} 个组件",
        f"漏洞统计：共 {len(vulnerabilities)} 个漏洞，严重 {counts['critical']}，高危 {counts['high']}，中危 {counts['medium']}，低危 {counts['low']}",
        "二、漏洞可信度说明",
        f"已确认漏洞：{confidence['confirmed']}；待确认漏洞：{confidence['review']}；疑似误报漏洞：{confidence['false_positive']}。报告统计默认以已确认漏洞为主，待确认和疑似误报单独列示，避免误报影响管理层判断。",
        "三、整改优先级清单",
        "；".join(priority_lines) if priority_lines else "暂无需要整改的已确认漏洞。",
        "四、开发修复建议",
        "；".join(fix_lines) if fix_lines else "暂无开发侧升级命令建议。",
        "五、多工具扫描结果汇总",
        "本平台采用多工具联合分析机制，综合 OpenSCA、Syft、Trivy 与 OWASP Dependency-Track 的结果，对开源组件、SBOM、漏洞、License 和风险指标进行标准化、去重、合并和可信度评分。最终结果不是单一扫描器输出，而是经过多源交叉验证后的统一风险视图。",
        f"OpenSCA 成功任务数：{scan_tasks.get('opensca', 0)}；Syft 成功任务数：{scan_tasks.get('syft', 0)}；Trivy 成功任务数：{scan_tasks.get('trivy', 0)}；Dependency-Track 成功任务数：{scan_tasks.get('dependency_track', 0)}；失败/超时任务数：{scan_tasks.get('failed', 0)}；跳过任务数：{scan_tasks.get('skipped', 0)}。",
        "工具状态明细：" + _tool_status_text(scan_tasks.get("details")),
        f"高可信组件：{component_confidence['high']}；中可信组件：{component_confidence['medium']}；低可信组件：{component_confidence['low']}；待确认组件：{component_confidence['review']}。",
        "六、扫描完整性与识别可信度说明",
        f"是否启用兜底识别：{'是' if fallback_enabled else '否'}；扫描模式：{', '.join(modes) or 'unknown'}；版本未知组件：{component_confidence['unknown_version']}；待人工确认组件：{component_confidence['manual_confirm']}。",
        "本次扫描未发现完整依赖清单文件时，平台会启用兜底识别模式。由于缺少标准依赖声明或锁定文件，部分组件版本可能无法精确确认，相关漏洞结果已按可信度进行区分，低可信和版本未知结果不直接纳入高危漏洞结论。",
        "依赖版本未锁定风险：" + ("；".join(unlocked_lines) if unlocked_lines else "暂无未锁定版本、版本范围或版本缺失风险。"),
        "漏洞统计图：critical/high/medium/low = "
        + "/".join(str(counts[key]) for key in ("critical", "high", "medium", "low")),
        "高危漏洞：" + ("；".join(f"{item.cve_id or item.advisory_id} {item.package_name} {item.cvss_score}" for item in high_risk[:10]) or "暂无"),
        "修复建议：优先升级存在严重、高危漏洞的组件；无法升级时采用隔离、最小权限和访问控制降低暴露面。",
        "风险趋势：按漏洞发布时间持续跟踪新增漏洞，建议每次源码变更或镜像发布前重新扫描。",
        "等保整改建议：建立组件台账、漏洞处置闭环、补丁验证记录和供应链安全审计证据。",
    ]
    return lines


def _xml_text(value: object) -> str:
    text = "" if value is None else str(value)
    return html.escape(text.replace("\n", " / "))


PAGE_WIDTH_DXA = 11906
PAGE_HEIGHT_DXA = 16838
PAGE_MARGIN_DXA = 1440
CONTENT_WIDTH_DXA = PAGE_WIDTH_DXA - PAGE_MARGIN_DXA * 2


def _run(text: object, *, bold: bool = False, size: int | None = None, color: str | None = None) -> str:
    props = []
    if bold:
        props.append("<w:b/>")
    if size:
        props.append(f'<w:sz w:val="{size}"/>')
    if color:
        props.append(f'<w:color w:val="{color}"/>')
    properties = f"<w:rPr>{''.join(props)}</w:rPr>" if props else ""
    space = ' xml:space="preserve"' if str(text).strip() != str(text) else ""
    return f"<w:r>{properties}<w:t{space}>{_xml_text(text)}</w:t></w:r>"


def _paragraph(
    text: object = "",
    *,
    style: str | None = None,
    align: str | None = None,
    bold: bool = False,
    size: int | None = None,
    color: str | None = None,
    page_break: bool = False,
    before: int | None = None,
    after: int | None = None,
    line: int | None = None,
) -> str:
    p_props = []
    if style:
        p_props.append(f'<w:pStyle w:val="{style}"/>')
    if align:
        p_props.append(f'<w:jc w:val="{align}"/>')
    if before is not None or after is not None or line is not None:
        spacing = []
        if before is not None:
            spacing.append(f'w:before="{before}"')
        if after is not None:
            spacing.append(f'w:after="{after}"')
        if line is not None:
            spacing.append(f'w:line="{line}" w:lineRule="auto"')
        p_props.append(f"<w:spacing {' '.join(spacing)}/>")
    properties = f"<w:pPr>{''.join(p_props)}</w:pPr>" if p_props else ""
    content = _run(text, bold=bold, size=size, color=color) if text != "" else ""
    if page_break:
        content += '<w:r><w:br w:type="page"/></w:r>'
    return f"<w:p>{properties}{content}</w:p>"


def _page_break() -> str:
    return _paragraph(page_break=True)


def _cell_spec(value: object, **overrides: object) -> dict[str, object]:
    data = {"value": value}
    data.update(overrides)
    return data


def _cell_paragraphs(value: object, *, align: str | None, bold: bool, size: int | None = None) -> str:
    if isinstance(value, (list, tuple)):
        values = list(value)
    else:
        values = str(value or "").splitlines() or [""]
    return "".join(_paragraph(item, align=align, bold=bold, size=size, after=0, line=300) for item in values)


def _table_cell(
    value: object,
    *,
    header: bool = False,
    cols: int = 1,
    width: int | None = None,
    align: str | None = None,
    fill: str | None = None,
    bold: bool | None = None,
    size: int | None = None,
) -> str:
    span = f'<w:gridSpan w:val="{cols}"/>' if cols > 1 else ""
    shading = f'<w:shd w:fill="{fill or "E5E7EB"}"/>' if header or fill else ""
    cell_width = f'<w:tcW w:w="{width}" w:type="dxa"/>' if width else ""
    props = (
        f"<w:tcPr>{cell_width}{span}{shading}<w:vAlign w:val=\"center\"/>"
        '<w:tcMar><w:top w:w="110" w:type="dxa"/><w:left w:w="140" w:type="dxa"/>'
        '<w:bottom w:w="110" w:type="dxa"/><w:right w:w="140" w:type="dxa"/></w:tcMar></w:tcPr>'
    )
    content = _cell_paragraphs(value, align=align, bold=header if bold is None else bold, size=size)
    return f"<w:tc>{props}{content}</w:tc>"


def _table(
    rows: list[list[object]],
    *,
    header_rows: int = 1,
    widths: list[int] | None = None,
    border_color: str = "CBD5E1",
    border_size: int = 4,
    table_width: int = CONTENT_WIDTH_DXA,
) -> str:
    if not rows:
        return ""
    column_count = len(widths) if widths else max(len(row) for row in rows)
    if not widths:
        base = table_width // column_count
        widths = [base] * column_count
        widths[-1] += table_width - sum(widths)
    body = []
    for index, row in enumerate(rows):
        grid_index = 0
        cell_xml = []
        for value in row:
            if isinstance(value, dict):
                cell_value = value.get("value", "")
                cols = int(value.get("cols") or 1)
                header = bool(value.get("header", index < header_rows))
                align = str(value.get("align") or ("center" if header else "left"))
                fill = str(value.get("fill") or "") or None
                bold = bool(value.get("bold")) if "bold" in value else None
                size = int(value["size"]) if value.get("size") else None
            else:
                cell_value = value
                cols = 1
                header = index < header_rows
                align = "center" if header else "left"
                fill = None
                bold = None
                size = None
            width = sum(widths[grid_index : grid_index + cols])
            cell_xml.append(_table_cell(cell_value, header=header, cols=cols, width=width, align=align, fill=fill, bold=bold, size=size))
            grid_index += cols
        cells = "".join(cell_xml)
        body.append(f"<w:tr>{cells}</w:tr>")
    borders = (
        f'<w:tblBorders><w:top w:val="single" w:sz="{border_size}" w:color="{border_color}"/>'
        f'<w:left w:val="single" w:sz="{border_size}" w:color="{border_color}"/>'
        f'<w:bottom w:val="single" w:sz="{border_size}" w:color="{border_color}"/>'
        f'<w:right w:val="single" w:sz="{border_size}" w:color="{border_color}"/>'
        f'<w:insideH w:val="single" w:sz="{border_size}" w:color="{border_color}"/>'
        f'<w:insideV w:val="single" w:sz="{border_size}" w:color="{border_color}"/></w:tblBorders>'
    )
    props = f'<w:tblPr><w:tblW w:w="{sum(widths)}" w:type="dxa"/><w:tblLayout w:type="fixed"/>{borders}</w:tblPr>'
    grid = f"<w:tblGrid>{''.join(f'<w:gridCol w:w=\"{width}\"/>' for width in widths)}</w:tblGrid>"
    return f"<w:tbl>{props}{grid}{''.join(body)}</w:tbl>"


def _severity_label(value: str | None) -> str:
    return SEVERITY_LABELS.get(value or "unknown", value or "未知风险")


def _date_only(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text[:10].replace("/", "-")


def _component_group_id(component: Component | None, vulnerability: VulnerabilityRecord | None = None) -> str:
    if component:
        return component.group_id or component.normalized_name or component.package_name
    return vulnerability.package_name if vulnerability else ""


def _component_type(component: Component) -> str:
    return "开源组件"


def _dependency_label(value: str) -> str:
    text = (value or "").lower()
    if text in {"indirect", "transitive"}:
        return "间接引入"
    if text == "base_image":
        return "基础镜像"
    return "直接引入"


def _version_date(snapshot: RiskMonitorSnapshot | None) -> str:
    return snapshot.current_version_published_at if snapshot and snapshot.current_version_published_at else "未知"


def _component_age(snapshot: RiskMonitorSnapshot | None) -> str:
    if snapshot and snapshot.component_age_years:
        return f"{snapshot.component_age_years:.1f}年"
    return "未知"


def _component_activity(snapshot: RiskMonitorSnapshot | None) -> tuple[str, str]:
    if not snapshot:
        return ("待确认", "未执行版本监测，缺少版本发布日期、最新版本和维护状态信息。")
    age = float(snapshot.component_age_years or 0)
    eol_status = str(snapshot.eol_status or "").strip().lower()
    eol_values = {"eol", "end_of_life", "deprecated", "unsupported", "retired", "abandoned", "inactive", "已停更", "停止维护"}
    if eol_status in eol_values:
        activity = "已停更"
    elif snapshot.update_available:
        activity = "活跃"
    elif eol_status in {"review", "suspected", "待确认", "需复核"}:
        activity = "待确认"
    elif age >= 5:
        activity = "已停更"
    elif age >= 3:
        activity = "低活跃"
    else:
        activity = "正常"

    details = []
    if snapshot.current_version_published_at:
        details.append(f"版本发布日期：{snapshot.current_version_published_at}")
    if age:
        details.append(f"组件年龄：{age:.1f}年")
    if snapshot.latest_version:
        details.append(f"最新版本：{snapshot.latest_version}")
    if snapshot.update_available:
        details.append(f"存在可升级版本，版本差异：{snapshot.version_delta or '待确认'}")
    if eol_status in {"review", "suspected", "待确认", "需复核"}:
        details.append("生命周期状态需复核")
    if snapshot.recommendation:
        details.append(snapshot.recommendation)
    if not details:
        details.append("版本监测未返回可用于判断活跃度的详细信息。")
    return activity, "；".join(details)


def _is_latest(snapshot: RiskMonitorSnapshot | None) -> str:
    if not snapshot or not snapshot.latest_version:
        return "未执行版本监测" if not snapshot else "未获取到最新版本"
    return "否" if snapshot.update_available else "是"


def _latest_version(snapshot: RiskMonitorSnapshot | None) -> str:
    if not snapshot:
        return "未执行版本监测"
    return snapshot.latest_version if snapshot.latest_version else "未获取到最新版本"


def _component_vulnerabilities(vulnerabilities: list[VulnerabilityRecord]) -> dict[int, list[VulnerabilityRecord]]:
    grouped: dict[int, list[VulnerabilityRecord]] = defaultdict(list)
    for item in _confirmed_vulnerabilities(vulnerabilities):
        if item.component_id:
            grouped[item.component_id].append(item)
    return grouped


def _component_risk_label(items: list[VulnerabilityRecord]) -> str:
    if not items:
        return "无漏洞"
    severities = {item.severity for item in items}
    if "critical" in severities:
        return "严重风险"
    if "high" in severities:
        return "高危风险"
    if "medium" in severities:
        return "中危风险"
    if "low" in severities:
        return "低危风险"
    return "未知风险"


def _vulnerability_distribution(items: list[VulnerabilityRecord]) -> str | int:
    if not items:
        return 0
    counts = Counter(item.severity if item.severity in SEVERITY_LABELS else "unknown" for item in items)
    return f"严重{counts['critical']}；高危{counts['high']}；中危{counts['medium']}；低危{counts['low']}"


def _recommended_version(component: Component, items: list[VulnerabilityRecord], snapshot: RiskMonitorSnapshot | None) -> str:
    fixed = next((item.fixed_version for item in items if item.fixed_version), "")
    if fixed:
        return fixed
    if snapshot and snapshot.latest_version:
        return snapshot.latest_version
    return "暂无推荐"


def _exploit_difficulty(item: VulnerabilityRecord) -> str:
    if item.exploited_in_wild or item.has_poc or item.cvss_score >= 7:
        return "容易"
    if item.cvss_score >= 4:
        return "一般"
    if item.cvss_score > 0:
        return "困难"
    return "未知"


def _vulnerability_reference_url(item: VulnerabilityRecord) -> str:
    if item.detail_url:
        return item.detail_url
    if item.advisory_id and item.source == "osv":
        return f"https://osv.dev/vulnerability/{item.advisory_id}"
    if item.cve_id:
        return f"https://nvd.nist.gov/vuln/detail/{item.cve_id}"
    return ""


def _vulnerability_description(item: VulnerabilityRecord) -> str:
    return item.description or item.priority_reason or item.business_impact or "暂无描述"


def _vulnerability_solution_reference(component: Component | None, item: VulnerabilityRecord) -> str:
    lines = [_fix_command(component, item)]
    if item.fixed_version:
        lines.append(f"修复版本：{item.fixed_version}")
    reference_url = _vulnerability_reference_url(item)
    if reference_url:
        lines.append(f"参考链接：{reference_url}")
    return "\n".join(lines)


def _vulnerability_extra_info(item: VulnerabilityRecord) -> str:
    details = []
    if item.priority_reason:
        details.append(f"优先级依据：{item.priority_reason}")
    if item.business_impact:
        details.append(f"业务影响：{item.business_impact}")
    if item.match_reason:
        details.append(f"匹配依据：{item.match_reason}")
    if item.version_range:
        details.append(f"影响版本范围：{item.version_range}")
    reachability_parts = []
    if item.reachability_status and item.reachability_status != "unknown":
        reachability_parts.append(item.reachability_status)
    if item.reachability_evidence:
        reachability_parts.append(item.reachability_evidence)
    if reachability_parts:
        details.append("可达性：" + "；".join(reachability_parts))
    if item.call_path_summary:
        details.append(f"调用路径：{item.call_path_summary}")
    return "；".join(details) if details else "暂无补充信息"


def _report_date() -> str:
    return datetime.now(timezone.utc).strftime("%Y年%m月%d日")


def _component_type_summary(components: list[Component]) -> str:
    ecosystems = sorted({item.ecosystem for item in components if item.ecosystem})
    return "、".join(ecosystems) if ecosystems else "开源组件"


def _source_files(project: Project, components: list[Component]) -> list[list[object]]:
    files = []
    seen = set()
    for component in components:
        source = component.source_file or component.source_path
        if source and source not in seen:
            seen.add(source)
            files.append([len(files) + 1, project.name, source])
    return files or [[1, project.name, project.scan_note or "系统源代码"]]


def _component_rows(components: list[Component], version_cache: dict[int, RiskMonitorSnapshot] | None = None, limit: int = 120) -> list[list[object]]:
    version_cache = version_cache or {}
    rows = [["序号", "组件名称", "当前版本", "组件类型", "许可协议", "识别可信度"]]
    rows.extend(
        [
            [
                index,
                component.package_name,
                _display_component_version(component, version_cache.get(component.id)),
                component.ecosystem or "unknown",
                _display_license(component.license_name),
                component.confidence_level or "High",
            ]
            for index, component in enumerate(components[:limit], start=1)
        ]
    )
    return rows


def _risk_rows(vulnerabilities: list[VulnerabilityRecord]) -> list[list[object]]:
    rows = [["风险等级", "漏洞数量"]]
    counts = _risk_counts(vulnerabilities)
    rows.extend(
        [
            ["严重风险", counts["critical"]],
            ["高危风险", counts["high"]],
            ["中危风险", counts["medium"]],
            ["低危风险", counts["low"]],
            ["未知风险", counts["unknown"]],
        ]
    )
    return rows


def _vulnerability_rows(vulnerabilities: list[VulnerabilityRecord], limit: int = 80) -> list[list[object]]:
    rows = [["序号", "组件名称", "组件版本", "风险等级", "CVE/公告", "修复版本", "备注"]]
    for index, vulnerability in enumerate(_priority_sorted(_confirmed_vulnerabilities(vulnerabilities))[:limit], start=1):
        rows.append(
            [
                index,
                vulnerability.package_name,
                vulnerability.package_version,
                _severity_label(vulnerability.severity),
                vulnerability.cve_id or vulnerability.advisory_id,
                vulnerability.fixed_version or "待确认",
                vulnerability.priority_reason or vulnerability.description or "已确认风险",
            ]
        )
    if len(rows) == 1:
        rows.append(["-", "暂无", "-", "暂无", "-", "-", "未发现已确认漏洞"])
    return rows


def _cover_date(value: str) -> str:
    match = re.search(r"(\d{4})年(\d{2})月(\d{2})日", value)
    if not match:
        return value
    year, month, day = match.groups()
    return f"{year} 年 {month} 月 {day} 日"


def _cover_page(project: Project, metadata: dict[str, str], date_text: str) -> list[str]:
    return [
        _paragraph(metadata["client_name"], align="center", bold=True, size=44, before=760, after=360),
        _paragraph(project.name, align="center", bold=True, size=40, after=1180),
        _paragraph("软件成分分析报告", align="center", bold=True, size=34, after=3300),
        _paragraph(_cover_date(date_text), align="center", size=32, after=2100),
        _paragraph(metadata["organization_name"], align="center", bold=True, size=26, after=300),
        _paragraph("版权所有  侵权必究", align="center", bold=True, size=24),
        _page_break(),
    ]


def _report_properties_table(project: Project, components: list[Component], metadata: dict[str, str]) -> str:
    label_fill = "F2F2F2"
    widths = [2050, 1500, 1400, 1600, 1000, 1476]
    audit_dates = f"{metadata['audit_start_date']} 至 {metadata['audit_end_date']}"
    rows = [
        [
            _cell_spec(["项目名称", "Project Name"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(f"{project.name} SCA 审计", cols=5, align="left", size=22),
        ],
        [
            _cell_spec(["系统名称", "Software Name"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(project.name, cols=2, align="left", size=22),
            _cell_spec(["版本号", "Version Number"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["version_number"], cols=2, align="center", size=22),
        ],
        [
            _cell_spec(["委托单位名称", "Client Name"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["client_name"], cols=5, align="left", size=22),
        ],
        [
            _cell_spec(["委托单位地址", "Client Address"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["client_address"], cols=5, align="left", size=22),
        ],
        [
            _cell_spec(["联系人姓名", "Contactor Name"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["contact_name"], align="left", size=22),
            _cell_spec(["联系电话", "Phone"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["contact_phone"], align="left", size=22),
            _cell_spec(["邮箱", "E-mail"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["contact_email"], align="left", size=20),
        ],
        [
            _cell_spec(["审计机构名称", "Organization", "Name"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["organization_name"], cols=5, align="left", size=22),
        ],
        [
            _cell_spec(["审计地点", "Audit Address"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["audit_address"], cols=5, align="left", size=22),
        ],
        [
            _cell_spec(["样品内容及数量", "Audit Sample"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(f"系统源代码、系统开源组件 [{len(components)}]", cols=3, align="left", size=21),
            _cell_spec("其他 无", cols=2, align="center", size=21),
        ],
        [
            _cell_spec(["代码接收日期", "Accepted Date"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(metadata["accepted_date"], align="center", size=21),
            _cell_spec(["审计日期", "Testing Date"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(audit_dates, cols=3, align="center", size=21),
        ],
        [
            _cell_spec(["审计标准", "Audit Standard"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec("●  以《GB/T 39412-2020 信息安全技术 代码安全审查规范》为主要依据。", cols=5, align="left", size=20),
        ],
        [
            _cell_spec(["参考文件", "Reference", "Document"], header=True, fill=label_fill, align="center", bold=True),
            _cell_spec(
                [
                    "●  GB/T 39412-2020 信息安全技术 代码安全审查规范",
                    "●  OWASP 代码安全审查指南",
                    "●  CWE 代码安全审查指南",
                    "●  ISO/IEC 27034 应用安全标准",
                    "●  Python/C/C++/Java/PHP/Javascript 安全编码规范",
                ],
                cols=5,
                align="left",
                size=20,
            ),
        ],
        [
            _cell_spec("需求依据", header=True, fill=label_fill, align="center", bold=True),
            _cell_spec("●  《代码安全审计实施方案》", cols=5, align="left", size=20),
        ],
    ]
    return _table(rows, header_rows=0, widths=widths, border_color="000000", border_size=8)


def _write_docx(
    path: Path,
    project: Project,
    components: list[Component],
    vulnerabilities: list[VulnerabilityRecord],
    scan_tasks: dict[str, object],
    metadata: dict[str, str],
    version_cache: dict[int, RiskMonitorSnapshot],
) -> None:
    confirmed = _confirmed_vulnerabilities(vulnerabilities)
    confidence = _confidence_groups(vulnerabilities)
    component_confidence = _component_confidence_groups(components, version_cache)
    priority_items = _priority_sorted(confirmed)
    component_by_id = {item.id: item for item in components}
    date_text = _report_date()
    toc_items = [
        "1\t项目概述\t1",
        "1.1\t审核目的\t1",
        "1.2\t审核依据\t1",
        "1.3\t审计范围\t1",
        "1.4\t审计方法\t2",
        "2\t测试目标\t3",
        "3\t测试标准\t3",
        "4\t系统基本情况\t4",
        "4.1\t系统基本情况\t4",
        "4.2\t审查组件清单\t4",
        "5\t人员安排\t5",
        "6\t组件提交活跃度统计\t5",
        "7\t风险组件最新版本\t6",
        "8\t版权许可协议风险提示\t7",
        "9\t组件安全审计结果汇总\t8",
        "10\t审计结论及建议\t9",
        "附录A\t组件安全审计缺陷详情\t10",
        "附录B\t组件安全审计清单\t12",
        "附录C\t安全编码规范要求\t13",
    ]
    paragraphs = [
        *_cover_page(project, metadata, date_text),
        _paragraph("声 明", style="Heading1", align="center"),
        _paragraph("本报告无审核人员和授权签字人签字无效；"),
        _paragraph("本报告涂改无效；"),
        _paragraph("未经委托单位书面批准，不得复制报告（完整复制除外）；"),
        _paragraph("本报告审计结果仅对委托单位当时提供的源代码、依赖清单、构建产物和开源组件信息有效。当被测代码发生变更后，报告结论需重新验证。"),
        _paragraph("本报告结论的有效性建立在委托单位提供材料真实性和扫描环境完整性的基础上。"),
        _paragraph("报告属性信息", align="center", bold=True, size=30, before=240, after=120),
        _paragraph("(Report Properties Information)", align="center", bold=True, size=25, after=120),
        _report_properties_table(project, components, metadata),
        _page_break(),
        _paragraph("目  录", style="Title"),
        *[_paragraph(item, style="TOC") for item in toc_items],
        _page_break(),
        _paragraph("1 项目概述", style="Heading1"),
        _paragraph("组件安全审计工作通过分析当前业务系统的源代码、依赖声明、构建文件和扫描结果，识别系统引入的开源组件、许可证与已公开漏洞，并给出风险等级、整改优先级和修复建议。"),
        _paragraph("1.1 审核目的", style="Heading2"),
        _paragraph("本次审计用于检查业务系统开源组件使用情况，发现存在安全漏洞、许可证风险、版本缺失或版本未锁定的组件，支撑后续整改和复测。"),
        _paragraph("1.2 审核依据", style="Heading2"),
        _paragraph("本次审计以 GB/T 39412-2020、CWE TOP25、NVD/CVE、OSV、GitHub Advisory 以及多工具扫描结果为参考依据。"),
        _paragraph("1.3 审计范围", style="Heading2"),
        _paragraph(f"本次审计范围为项目“{project.name}”提交的源代码、依赖文件、SBOM 信息和平台扫描结果。项目备注：{project.scan_note or '无'}。"),
        _paragraph("1.4 审计方法", style="Heading2"),
        _paragraph("审计采用工具审查、规则匹配、多源漏洞合并、可信度评分和人工复核标记相结合的方式，最终形成统一的组件安全风险视图。"),
        _table(
            [
                ["检测类型", "检测项", "描述"],
                ["漏洞检测", "CVE/OSV/GHSA", "识别开源组件公开漏洞与受影响版本范围"],
                ["许可证检测", "License", "识别组件许可证并提示潜在合规风险"],
                ["版本风险", "版本锁定/版本未知", "识别未锁定版本、版本范围和不可确认版本"],
                ["可信度分析", "多工具交叉验证", "区分已确认、待确认和疑似误报结果"],
            ]
        ),
        _paragraph("2 测试目标", style="Heading1"),
        _paragraph("确认系统开源组件清单、漏洞数量、风险等级、影响范围、修复版本和整改优先级，为上线评审、等保整改和供应链安全治理提供依据。"),
        _paragraph("3 测试标准", style="Heading1"),
        _paragraph("测试结果按照严重、高危、中危、低危和未知风险分级，并结合 P0/P1/P2/P3/Review 优先级、CVSS 评分、可信度和修复版本可用性确定整改顺序。"),
        _paragraph("4 系统基本情况", style="Heading1"),
        _paragraph("4.1 系统基本情况", style="Heading2"),
        _table([["编号", "基本项", "描述"], ["1", "组件数量", len(components)], ["2", "组件类型", _component_type_summary(components)], ["3", "漏洞记录", len(vulnerabilities)], ["4", "上线建议", _release_advice(confirmed)]]),
        _paragraph("4.2 审查组件清单", style="Heading2"),
        _table([["序号", "检测目标", "组件依赖文件包"], *_source_files(project, components)]),
        _paragraph("5 人员安排", style="Heading1"),
        _table([["编号", "参与人员", "负责内容"], ["1", metadata["auditor_name"], "组件识别、SBOM 生成、漏洞匹配、报告编制"], ["2", metadata["reviewer_name"], "风险确认、报告审查、整改建议"], ["3", metadata["quality_reviewer_name"], "质量审核"]]),
        _paragraph("6 组件提交活跃度统计", style="Heading1"),
        _table([["序号", "组件名称", "当前版本", "发布日期", "活跃度", "活跃度参考说明"], *[[index, item.package_name, _display_component_version(item, version_cache.get(item.id)), _version_date(version_cache.get(item.id)), "未执行版本监测" if not version_cache.get(item.id) else "正常", f"组件年龄：{_component_age(version_cache.get(item.id))}"] for index, item in enumerate(components[:20], start=1)]]),
        _paragraph("7 风险组件最新版本", style="Heading1"),
        _table([["序号", "组件名称", "当前版本", "最新版本", "当前版本是否最新版本"], *[[index, item.package_name, item.package_version, _latest_version(version_cache.get(item.component_id or 0)), _is_latest(version_cache.get(item.component_id or 0))] for index, item in enumerate(priority_items[:20], start=1)]] or [["序号", "组件名称", "当前版本", "最新版本", "当前版本是否最新版本"], ["-", "暂无", "-", "-", "待确认"]]),
        _paragraph("8 风险组件版本年龄统计", style="Heading1"),
        _table([["序号", "组件名称", "当前版本", "发布日期", "组件版本年龄"], *[[index, item.package_name, _display_component_version(item, version_cache.get(item.id)), _version_date(version_cache.get(item.id)), _component_age(version_cache.get(item.id))] for index, item in enumerate(components[:20], start=1)]]),
        _paragraph("9 版权许可协议风险提示", style="Heading1"),
        _paragraph("许可证风险需结合组件使用方式、分发方式、修改情况和企业合规要求进行复核。下表列出本次识别到的组件许可证信息。"),
        _table([["许可协议简称", "许可协议全称", "风险说明", "使用范围", "版权/使用条件", "使用限制", "是否兼容GPL"], *[[license_policy(name).short_name if license_policy(name).short_name != "待确认" else name, license_policy(name).full_name, license_policy(name).risk_note, license_policy(name).scope, license_policy(name).conditions, license_policy(name).limitations, license_policy(name).gpl_compatible] for name in sorted({_display_license(item.license_name) for item in components})]]),
        _table([["序号", "组件名称", "当前版本", "许可协议"], *[[index, item.package_name, _display_component_version(item, version_cache.get(item.id)), _display_license(item.license_name)] for index, item in enumerate(components[:80], start=1)]]),
        _paragraph("10 组件安全审计结果汇总", style="Heading1"),
        _paragraph("本次扫描结论摘要", style="Heading2"),
        _paragraph(_natural_summary(project, components, vulnerabilities)),
        _table(_risk_rows(vulnerabilities)),
        _paragraph("漏洞可信度说明", style="Heading2"),
        _paragraph(f"已确认漏洞：{confidence['confirmed']}；待确认漏洞：{confidence['review']}；疑似误报漏洞：{confidence['false_positive']}。报告统计默认以已确认漏洞为主，待确认和疑似误报单独列示。"),
        _paragraph("多工具扫描结果汇总", style="Heading2"),
        _table(
            [
                ["扫描引擎", "成功任务数"],
                ["OpenSCA", scan_tasks.get("opensca", 0)],
                ["Syft", scan_tasks.get("syft", 0)],
                ["Trivy", scan_tasks.get("trivy", 0)],
                ["Dependency-Track", scan_tasks.get("dependency_track", 0)],
                ["失败/超时", scan_tasks.get("failed", 0)],
            ]
        ),
        _paragraph("工具状态明细", style="Heading2"),
        _table(_tool_status_rows(scan_tasks)),
        _paragraph("组件安全测试", style="Heading2"),
        _table(_vulnerability_rows(vulnerabilities)),
        _paragraph("11 审计结论及建议", style="Heading1"),
        _paragraph(f"上线建议：{_release_advice(confirmed)}"),
        _paragraph("整改优先级清单", style="Heading2"),
        _table([["优先级", "CVE/公告", "组件", "当前版本", "等级", "风险分", "修复版本", "建议期限"], *[[item.risk_priority or "Review", item.cve_id or item.advisory_id, item.package_name, item.package_version, _severity_label(item.severity), item.risk_score, item.fixed_version or "安全版本", item.suggested_deadline or "按计划修复"] for item in priority_items[:20]]] or [["优先级", "CVE/公告", "组件", "当前版本", "等级", "风险分", "修复版本", "建议期限"], ["-", "暂无", "-", "-", "-", "-", "-", "-"]]),
        _paragraph("开发修复建议", style="Heading2"),
        *[_paragraph(_fix_command(component_by_id.get(item.component_id or 0), item)) for item in priority_items[:12]],
        _paragraph("附录A 组件安全审计缺陷详情", style="Heading1"),
    ]
    if priority_items:
        for item in priority_items[:20]:
            paragraphs.extend(
                [
                    _table([["组件名称", item.package_name], ["当前版本", item.package_version], ["风险等级", _severity_label(item.severity)], ["检出时间", item.published_at_text or date_text], ["版本日期", _version_date(version_cache.get(item.component_id or 0))], ["组件年龄", _component_age(version_cache.get(item.component_id or 0))]], header_rows=0),
                    _table(
                        [
                            ["漏洞名称", item.advisory_id or item.cve_id or "组件漏洞"],
                            ["风险等级", _severity_label(item.severity)],
                            ["CVE编号", item.cve_id or item.advisory_id or "待确认"],
                            ["CWE编号", item.cwe_id or "-"],
                            ["受影响组件", f"{item.package_name} {item.package_version}"],
                            ["发布时间", item.published_at_text or "未知"],
                            ["漏洞描述", item.description or item.priority_reason or "暂无描述"],
                            ["修复建议", _fix_command(component_by_id.get(item.component_id or 0), item)],
                        ],
                        header_rows=0,
                    ),
                ]
            )
    else:
        paragraphs.append(_paragraph("本次审计未发现已确认组件安全缺陷。"))
    paragraphs.extend(
        [
            _paragraph("附录B 组件安全审计清单", style="Heading1"),
            _paragraph(f"高可信组件：{component_confidence['high']}；中可信组件：{component_confidence['medium']}；低可信组件：{component_confidence['low']}；待确认组件：{component_confidence['review']}。"),
            _table(_component_rows(components, version_cache)),
            _paragraph("附录C 安全编码规范要求", style="Heading1"),
            _table(
                [
                    ["类别", "规范要求"],
                    ["输入验证", "对外部输入进行白名单校验，避免注入和越权访问。"],
                    ["身份认证", "统一认证入口，密码和凭据不得硬编码在源码中。"],
                    ["访问控制", "按最小权限原则限制敏感接口、数据和后台任务。"],
                    ["日志审计", "记录关键安全事件，避免日志输出敏感信息。"],
                    ["组件治理", "建立组件台账、版本锁定、漏洞响应和复测闭环。"],
                ]
            ),
            _paragraph("附录D 审计人员详细信息", style="Heading1"),
            _table([["姓    名", metadata["auditor_name"]], ["角色", "人工审计，撰写报告"], ["复核人员", metadata["reviewer_name"]], ["质量审核", metadata["quality_reviewer_name"]]], header_rows=0),
        ]
    )
    section = '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/><w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="720" w:footer="720" w:gutter="0"/></w:sectPr>'
    document = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>{''.join(paragraphs)}{section}</w:body></w:document>"""
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", CONTENT_TYPES)
        archive.writestr("_rels/.rels", WORD_RELS)
        archive.writestr("word/styles.xml", WORD_STYLES)
        archive.writestr("word/document.xml", document)


def _style_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(color="FFFFFF", bold=True)
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if cell.row == 1:
                    cell.fill = title_fill
                    cell.font = title_font
                elif cell.row == 2 or all(sheet.cell(cell.row, index).value is not None for index in range(1, min(sheet.max_column, 4) + 1)):
                    if cell.row in {2, 11} or cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
        for column_index in range(1, sheet.max_column + 1):
            max_length = 10
            for cell in sheet[get_column_letter(column_index)]:
                max_length = max(max_length, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 42)
        sheet.freeze_panes = "A2"


def _write_xlsx_report(
    path: Path,
    project: Project,
    components: list[Component],
    vulnerabilities: list[VulnerabilityRecord],
    metadata: dict[str, str],
    version_cache: dict[int, RiskMonitorSnapshot],
    upload_records: list[UploadFileRecord],
) -> None:
    workbook = Workbook()
    task_sheet = workbook.active
    task_sheet.title = "任务信息"
    latest_upload = upload_records[0] if upload_records else None
    task_sheet.append(["任务信息", None])
    task_sheet.append(["项目名称", project.name])
    task_sheet.append(["任务文件信息", latest_upload.original_filename if latest_upload else project.scan_note or "系统源代码"])
    task_sheet.append(["特征文件", components[0].package_name if components else "待确认"])
    task_sheet.append(["任务标签", components[0].source_file or components[0].source_path if components else project.scan_note])
    task_sheet.append(["审计人员", metadata["auditor_name"]])
    task_sheet.append(["创建时间", metadata["accepted_date"]])
    task_sheet.append(["完成时间", metadata["audit_end_date"]])

    confirmed = _confirmed_vulnerabilities(vulnerabilities)
    vulnerabilities_by_component = _component_vulnerabilities(vulnerabilities)
    component_risks = [_component_risk_label(vulnerabilities_by_component.get(component.id, [])) for component in components]
    overview = workbook.create_sheet("审计概览")
    component_counts = Counter(component_risks)
    vuln_counts = _risk_counts(vulnerabilities)
    overview_rows = [
        ["组件安全分布", None],
        ["组件等级", "数量"],
        ["高危组件", component_counts["高危风险"] + component_counts["严重风险"]],
        ["中危组件", component_counts["中危风险"]],
        ["低危组件", component_counts["低危风险"]],
        ["未知风险组件", component_counts["未知风险"]],
        ["未知版本组件", _component_unknown_version_count(components, version_cache)],
        ["无漏洞组件", component_counts["无漏洞"]],
        [],
        ["漏洞等级分布", None],
        ["组件等级", "数量"],
        ["严重漏洞", vuln_counts["critical"]],
        ["高危漏洞", vuln_counts["high"]],
        ["中危漏洞", vuln_counts["medium"]],
        ["低危漏洞", vuln_counts["low"]],
        ["未知漏洞", vuln_counts["unknown"]],
        [],
        ["基础统计", None],
        ["组件数量", len(components)],
        ["漏洞数量", len(vulnerabilities)],
        ["已确认漏洞", len(confirmed)],
    ]
    for row in overview_rows:
        overview.append(row)

    asset_sheet = workbook.create_sheet("审计资产列表")
    asset_sheet.append(["组件", "版本", "是否最新版本", "最新版本", "组ID", "组件类型", "组件路径", "许可证", "风险等级", "漏洞分布", "依赖关系", "推荐版本"])
    for component in components:
        items = vulnerabilities_by_component.get(component.id, [])
        snapshot = version_cache.get(component.id)
        asset_sheet.append(
            [
                component.package_name,
                _display_component_version(component, snapshot),
                _is_latest(snapshot),
                _latest_version(snapshot),
                component.group_id or component.normalized_name or component.package_name,
                _component_type(component),
                component.source_file or component.source_path,
                _display_license(component.license_name),
                _component_risk_label(items),
                _vulnerability_distribution(items),
                _dependency_label(component.dependency_type),
                _recommended_version(component, items, snapshot),
            ]
        )

    vulnerability_sheet = workbook.create_sheet("资产漏洞信息")
    vulnerability_sheet.append(
        [
            "漏洞编号",
            "严重程度",
            "发布日期",
            "CWE",
            "项目名",
            "组件",
            "版本",
            "漏洞利用难度",
            "创建日期",
            "组ID",
            "版本日期",
            "组件年龄",
            "活跃度",
            "活跃度说明",
            "漏洞描述",
            "解决方案参考",
            "补充信息",
            "确认状态",
            "可信度",
        ]
    )
    component_by_id = {component.id: component for component in components}
    for item in _priority_sorted(vulnerabilities):
        component = component_by_id.get(item.component_id or 0)
        snapshot = version_cache.get(item.component_id or 0)
        activity, activity_note = _component_activity(snapshot)
        vulnerability_sheet.append(
            [
                item.cve_id or item.advisory_id,
                _severity_label(item.severity),
                _date_only(item.published_at_text) or "未知",
                item.cwe_id or "-",
                f"{project.name} {project.scan_note or metadata['version_number']}".strip(),
                item.package_name,
                _vulnerability_component_version(component, item),
                _exploit_difficulty(item),
                metadata["audit_end_date"],
                _component_group_id(component, item),
                _version_date(snapshot),
                _component_age(snapshot),
                activity,
                activity_note,
                _vulnerability_description(item),
                _vulnerability_solution_reference(component, item),
                _vulnerability_extra_info(item),
                _vulnerability_review_status(item),
                f"{round(float(item.confidence_score or 0) * 100)}%",
            ]
        )

    license_sheet = workbook.create_sheet("许可协议信息")
    license_sheet.append(["许可协议信息", None])
    license_sheet.append(["许可协议简称", "许可协议全称", "风险说明", "使用范围", "使用条件", "使用限制", "是否兼容GPL", "OSI认证", "FSF认证", "风险等级", "许可描述", "组件数量", "涉及组件", "识别来源", "可信度"])
    for row in _license_rows(components):
        license_sheet.append(row)
    _style_workbook(workbook)
    workbook.save(path)


def _write_xlsx(path: Path, rows: list[list[str]]) -> None:
    row_xml = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for column_index, value in enumerate(row):
            column = chr(ord("A") + column_index)
            cells.append(f'<c r="{column}{row_index}" t="inlineStr"><is><t>{html.escape(str(value))}</t></is></c>')
        row_xml.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    sheet = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>{''.join(row_xml)}</sheetData></worksheet>"""
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", XLSX_TYPES)
        archive.writestr("_rels/.rels", XLSX_RELS)
        archive.writestr("xl/workbook.xml", WORKBOOK)
        archive.writestr("xl/_rels/workbook.xml.rels", WORKBOOK_RELS)
        archive.writestr("xl/worksheets/sheet1.xml", sheet)


def _write_pdf(path: Path, lines: list[str]) -> None:
    payload = json.dumps({"template": "聚信中文安全分析报告", "lines": lines}, ensure_ascii=False, indent=2)
    stream = f"BT /F1 12 Tf 50 760 Td ({payload[:900].replace('(', '[').replace(')', ']')}) Tj ET"
    body = f"""%PDF-1.4
1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj
2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj
3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >> endobj
4 0 obj << /Length {len(stream.encode("utf-8"))} >> stream
{stream}
endstream endobj
5 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj
trailer << /Root 1 0 R >>
%%EOF
"""
    path.write_bytes(body.encode("utf-8"))


def generate_report(db: Session, project_id: int, fmt: str, report_root: str, metadata: dict[str, object] | None = None) -> Path:
    project, components, vulnerabilities = _project_data(db, project_id)
    scan_tasks = _scan_task_summary(db, project_id)
    report_metadata = _report_metadata(metadata)
    version_cache = _latest_version_cache(db, project_id)
    uploads = _upload_records(db, project_id)
    output_dir = Path(report_root)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    path = output_dir / f"juxin-sca-report-{project_id}-{timestamp}.{fmt}"
    lines = _report_lines(project, components, vulnerabilities, scan_tasks, version_cache)
    if fmt == "docx":
        _write_docx(path, project, components, vulnerabilities, scan_tasks, report_metadata, version_cache)
    elif fmt == "xlsx":
        _write_xlsx_report(path, project, components, vulnerabilities, report_metadata, version_cache, uploads)
    elif fmt == "pdf":
        _write_pdf(path, lines)
    else:
        raise ValueError("不支持的报告格式")
    return path
